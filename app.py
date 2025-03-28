from flask import Flask, render_template, request
import openpyxl
import os

app = Flask(__name__)

EXCEL_FILE = "data.xlsx"

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        name = request.form['nm']
        email = request.form['em']
        number = request.form['no']
        date = request.form['date']
        ticket = request.form['tn']

        # Save form data to Excel
        save_to_excel(name, email, number, date, ticket)

    # Fetch queue data (only names)
    queue_data = read_from_excel()

    return render_template('index.html', queue_data=queue_data)

def save_to_excel(n, e, nu, d, t):
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Data"
        ws.append(["Name", "Email", "Number", "Date", "Ticket"])  # Header row
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([n, e, nu, d, t])  # Append new row
    wb.save(EXCEL_FILE)


def read_from_excel():
    if not os.path.exists(EXCEL_FILE):
        return []

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    return [[row[0]] for row in ws.iter_rows(min_row=2, values_only=True)]  # Only names

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)



# from flask import Flask, render_template

# app = Flask(__name__)

# @app.route("/")
# def home():
#     return render_template("index.html")

# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=5000)



